# -*- coding=utf-8 -*-
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

s = requests.session()
header = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36"
}

rs = s.get(url="http://www.xicidaili.com/nn/", headers=header)

soup = BeautifulSoup(rs.text, "lxml")
ip_list_all = []
ip_list = soup.select_one("#ip_list").select("tr")
ip_info_list_key = ["ip", "port", "address", "hidden", "type", "speed", "conn_time", "survival_time", "verify_time"]

for item in ip_list[1:]:
    ip_info_list_value = []
    ip_info = item.select("td")
    for info in ip_info[1:]:
        if info.select_one(".bar"):
            ip_info_list_value.append(info.select_one(".bar")["title"])
        else:
            ip_info_list_value.append(info.get_text().strip())
    ip_list_all.append(dict(zip(ip_info_list_key, ip_info_list_value)))

print(len(ip_list_all))

# 写excel文件
w = Workbook()  # 创建一个工作簿
ws = w.active
ws.title  = "西刺免费代理IP"  # 创建一个工作表
ws["A1"] = u"序号"
ws["B1"] = u"IP地址"
ws["C1"] = u"端口"
ws["D1"] = u"服务器地址"
ws["E1"] = u"是否匿名"
ws["F1"] = u"类型"
ws["G1"] = u"速度"
ws["H1"] = u"连接时间"
ws["I1"] = u"存活时间"
ws["J1"] = u"验证时间"
i = 0
for item in ip_list_all:
    i += 1
    ws["A%d" % i] = i  # 在i+1行1列写入
    ws["B%d" % i] = item["ip"]
    ws["C%d" % i] = item["port"]
    ws["D%d" % i] = item["address"]
    ws["E%d" % i] = item["hidden"]
    ws["F%d" % i] = item["type"]
    ws["G%d" % i] = item["speed"]
    ws["H%d" % i] = item["conn_time"]
    ws["I%d" % i] = item["survival_time"]
    ws["J%d" % i] = item["verify_time"]
    
w.save(u"西刺免费代理IP.xlsx")  # 保存

print("写excel完成")
